#!/usr/bin/env python3
"""
etl.py — Excel (Cape_P1_s.xlsx / Cape_P2_s.xlsx) → Cloudflare D1 SQL

Usage:
    python etl.py --p1 Cape_P1_s.xlsx --p2 Cape_P2_s.xlsx [--out load_questions.sql]

The script is idempotent: it uses INSERT ... ON CONFLICT DO UPDATE (upsert) on
the natural key (source_workbook, subject_id, month, year, paper, number, part,
subpart) so re-running after a spreadsheet fix updates rows in place without
creating duplicates.

Load the output into D1:
    wrangler d1 execute cape-questions --local --file=load_questions.sql  # local
    wrangler d1 execute cape-questions --file=load_questions.sql          # prod
"""

import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")


# =============================================================================
# Utility helpers
# =============================================================================

def norm(v):
    """Normalise a cell value: strip whitespace from strings, pass None/numbers through."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def sql_str(v) -> str:
    """Return a SQL literal for a value (NULL or a single-quoted, escaped string/number)."""
    if v is None:
        return "NULL"
    if isinstance(v, (int, float)):
        return str(v)
    # Escape single quotes by doubling them
    escaped = str(v).replace("'", "''")
    return f"'{escaped}'"


def build_diagram_key(subject: str, month, year, paper: int, number: int,
                      choice_label: str | None = None) -> str:
    """
    Deterministic diagram key.
    cape_{subject_slug}_{month_slug}_{year}_{paper}_{number}[_{choice_label}]

    Example: cape_accountingu1_may_2017_1_1
             cape_accountingu1_may_2017_1_1_a
    """
    subject_slug = re.sub(r'\s+', '', subject.lower())
    month_slug   = (str(month).lower() if month else 'unknown')
    year_val     = int(year) if year else 0
    base = f"cape_{subject_slug}_{month_slug}_{year_val}_{paper}_{number}"
    if choice_label:
        return f"{base}_{choice_label.lower()}"
    return base


# =============================================================================
# Row parsers
# =============================================================================

def get_header_index(ws) -> dict[str, int]:
    """Return {header_name: column_index} from the first row of the sheet."""
    headers = [norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    return {h: i for i, h in enumerate(headers) if h is not None}


def rows_from_p1(ws, sheet_name: str):
    """
    Yield one dict per non-blank Paper 1 question row.
    Includes 4 answer choices (A–D) with their own diagram keys.
    """
    idx = get_header_index(ws)
    if 'Question' not in idx:
        print(f"  [SKIP] Sheet '{sheet_name}' has no 'Question' column — skipping.")
        return

    for row in ws.iter_rows(min_row=2, values_only=True):
        q_text = norm(row[idx['Question']]) if 'Question' in idx else None
        if not q_text:
            continue  # blank template row

        subject = norm(row[idx['Subject']]) if 'Subject' in idx else None
        exam    = norm(row[idx['Exam']])    if 'Exam'    in idx else 'CAPE'
        month   = norm(row[idx['Month']])   if 'Month'   in idx else None
        year    = norm(row[idx['Year']])    if 'Year'    in idx else None
        paper   = norm(row[idx['Paper']])   if 'Paper'   in idx else 1
        number  = norm(row[idx['Number']])  if 'Number'  in idx else None
        section = norm(row[idx.get('Section', -1)]) if idx.get('Section') is not None else None
        topic   = norm(row[idx.get('Topic', -1)])   if idx.get('Topic') is not None else None
        diff    = norm(row[idx.get('Difficulty', -1)]) if idx.get('Difficulty') is not None else None
        correct = norm(row[idx.get('Correct', -1)])    if idx.get('Correct') is not None else None
        q_diag  = norm(row[idx.get('Question Diagram Path Prefix', -1)]) \
                  if idx.get('Question Diagram Path Prefix') is not None else None
        diag_present = norm(row[idx.get('Diagram Present', -1)]) \
                       if idx.get('Diagram Present') is not None else None
        q_code  = norm(row[idx.get('Validated Question Code', -1)]) \
                  if idx.get('Validated Question Code') is not None else q_text

        choices = []
        # Detect validated answer column name pattern once per sheet
        # Tries: "Validated Answer A", "Validated A", "Answer A Code", falling back to Answer A
        def validated_ans_key(lbl: str) -> str | None:
            for pat in (f'Validated Answer {lbl}', f'Validated {lbl}', f'Answer {lbl} Code'):
                if pat in idx:
                    return pat
            return None

        validated_warned = False
        for label in 'ABCD':
            ans_key      = f'Answer {label}'
            val_key      = validated_ans_key(label)
            diag_key_col = f'{label} Diagram Path Prefix'

            ans_text = norm(row[idx[ans_key]]) if ans_key in idx else None
            if not ans_text:
                continue

            ans_code = norm(row[idx[val_key]]) if val_key else None
            if ans_code is None and not validated_warned:
                print(f"    [WARN] No validated answer column found for '{label}' in '{sheet_name}' "
                      f"— using answer_raw as fallback. Check column names.")
                validated_warned = True
            ans_code = ans_code or ans_text  # fallback

            ans_diag = norm(row[idx[diag_key_col]]) if diag_key_col in idx else None
            choices.append({
                'label':       label,
                'answer_raw':  ans_text,
                'answer_code': ans_code,
                'diagram_key': ans_diag or build_diagram_key(
                    subject or '', month, year, int(paper or 1), int(number or 0), label
                ),
            })

        yield {
            'source_sheet': sheet_name,
            'exam':     str(exam or 'CAPE'),
            'subject':  str(subject or ''),
            'month':    month,
            'year':     int(year) if year is not None else None,
            'paper':    int(paper) if paper is not None else 1,
            'number':   int(number) if number is not None else 0,
            'part':     None,
            'subpart':  None,
            'section':  section,
            'topic':    topic,
            'difficulty': diff,
            'marks':    None,
            'correct_choice': correct,
            'question_raw':  q_text,
            'question_code': str(q_code or q_text),
            'question_diagram_key': q_diag or build_diagram_key(
                subject or '', month, year, int(paper or 1), int(number or 0)
            ),
            'diagram_present': 1 if diag_present and str(diag_present).strip().lower() in ('yes', '1', 'true') else 0,
            'choices': choices,
        }


def rows_from_p2(ws, sheet_name: str):
    """
    Yield one dict per non-blank Paper 2 question row.
    No choices; includes Part, Subpart, Marks.
    """
    idx = get_header_index(ws)
    if 'Question' not in idx:
        print(f"  [SKIP] Sheet '{sheet_name}' has no 'Question' column — skipping.")
        return

    for row in ws.iter_rows(min_row=2, values_only=True):
        q_text = norm(row[idx['Question']]) if 'Question' in idx else None
        if not q_text:
            continue

        subject = norm(row[idx['Subject']]) if 'Subject' in idx else None
        exam    = norm(row[idx['Exam']])    if 'Exam'    in idx else 'CAPE'
        month   = norm(row[idx['Month']])   if 'Month'   in idx else None
        year    = norm(row[idx['Year']])    if 'Year'    in idx else None
        paper   = norm(row[idx['Paper']])   if 'Paper'   in idx else 2
        number  = norm(row[idx['Number']])  if 'Number'  in idx else None
        part    = norm(row[idx.get('Part',    -1)]) if idx.get('Part')    is not None else None
        subpart = norm(row[idx.get('Subpart', -1)]) if idx.get('Subpart') is not None else None
        section = norm(row[idx.get('Section', -1)]) if idx.get('Section') is not None else None
        topic   = norm(row[idx.get('Topic',   -1)]) if idx.get('Topic')   is not None else None
        diff    = norm(row[idx.get('Difficulty', -1)]) if idx.get('Difficulty') is not None else None
        marks   = norm(row[idx.get('Marks', -1)])   if idx.get('Marks')   is not None else None
        q_diag  = norm(row[idx.get('Question Diagram Path Prefix', -1)]) \
                  if idx.get('Question Diagram Path Prefix') is not None else None
        diag_present = norm(row[idx.get('Diagram Present', -1)]) \
                       if idx.get('Diagram Present') is not None else None
        q_code  = norm(row[idx.get('Validated Question Code', -1)]) \
                  if idx.get('Validated Question Code') is not None else q_text

        yield {
            'source_sheet': sheet_name,
            'exam':     str(exam or 'CAPE'),
            'subject':  str(subject or ''),
            'month':    month,
            'year':     int(year) if year is not None else None,
            'paper':    int(paper) if paper is not None else 2,
            'number':   int(number) if number is not None else 0,
            'part':     str(part) if part is not None else None,
            'subpart':  str(subpart) if subpart is not None else None,
            'section':  section,
            'topic':    topic,
            'difficulty': diff,
            'marks':    float(marks) if marks is not None else None,
            'correct_choice': None,
            'question_raw':  q_text,
            'question_code': str(q_code or q_text),
            'question_diagram_key': q_diag or build_diagram_key(
                subject or '', month, year, int(paper or 2), int(number or 0)
            ),
            'diagram_present': 1 if diag_present and str(diag_present).strip().lower() in ('yes', '1', 'true') else 0,
            'choices': [],
        }


def rows_from_diagrams(ws, sheet_name: str):
    """
    Parse the Diagrams sheet and yield one dict per row with:
        diagram_key, drive_path, status
    """
    idx = get_header_index(ws)
    key_col  = next((k for k in idx if 'diagram' in k.lower() and 'key' in k.lower()), None)
    path_col = next((k for k in idx if 'corrected' in k.lower() or 'path' in k.lower()), None)
    stat_col = next((k for k in idx if 'status' in k.lower() or 'needs fix' in k.lower()), None)

    if not key_col or not path_col:
        print(f"  [WARN] Diagrams sheet '{sheet_name}': could not identify key/path columns — skipping.")
        return

    for row in ws.iter_rows(min_row=2, values_only=True):
        key  = norm(row[idx[key_col]])  if key_col  in idx else None
        path = norm(row[idx[path_col]]) if path_col in idx else None
        stat = norm(row[idx[stat_col]]) if stat_col in idx and stat_col else 'unknown'
        if not key or not path:
            continue
        yield {'diagram_key': key, 'drive_path': path, 'status': stat or 'unknown'}


# =============================================================================
# SQL emitters
# =============================================================================

def emit_subject_upsert(subject: str) -> str:
    return (
        f"INSERT INTO subjects (name) VALUES ({sql_str(subject)}) "
        f"ON CONFLICT(name) DO NOTHING;"
    )


def emit_question_upsert(row: dict, source_workbook: str) -> str:
    return f"""INSERT INTO questions (
  source_workbook, source_sheet, exam, subject_id, month, year, paper,
  number, part, subpart, section, topic, difficulty, marks, correct_choice,
  question_raw, question_code, question_diagram_key, diagram_present
) VALUES (
  {sql_str(source_workbook)},
  {sql_str(row['source_sheet'])},
  {sql_str(row['exam'])},
  (SELECT id FROM subjects WHERE name = {sql_str(row['subject'])}),
  {sql_str(row['month'])},
  {sql_str(row['year'])},
  {sql_str(row['paper'])},
  {sql_str(row['number'])},
  {sql_str(row['part'])},
  {sql_str(row['subpart'])},
  {sql_str(row['section'])},
  {sql_str(row['topic'])},
  {sql_str(row['difficulty'])},
  {sql_str(row['marks'])},
  {sql_str(row['correct_choice'])},
  {sql_str(row['question_raw'])},
  {sql_str(row['question_code'])},
  {sql_str(row['question_diagram_key'])},
  {sql_str(row['diagram_present'])}
)
ON CONFLICT(source_workbook, subject_id, month, year, paper, number, part, subpart)
DO UPDATE SET
  question_raw          = excluded.question_raw,
  question_code         = excluded.question_code,
  section               = excluded.section,
  topic                 = excluded.topic,
  difficulty            = excluded.difficulty,
  marks                 = excluded.marks,
  correct_choice        = excluded.correct_choice,
  question_diagram_key  = excluded.question_diagram_key,
  diagram_present       = excluded.diagram_present;"""


def emit_choice_upsert(row: dict, choice: dict, source_workbook: str) -> str:
    return f"""INSERT INTO choices (question_id, label, answer_raw, answer_code, diagram_key)
SELECT id, {sql_str(choice['label'])}, {sql_str(choice['answer_raw'])},
       {sql_str(choice['answer_code'])}, {sql_str(choice['diagram_key'])}
FROM questions
WHERE source_workbook = {sql_str(source_workbook)}
  AND subject_id = (SELECT id FROM subjects WHERE name = {sql_str(row['subject'])})
  AND month IS {sql_str(row['month'])}
  AND year  IS {sql_str(row['year'])}
  AND paper = {sql_str(row['paper'])}
  AND number = {sql_str(row['number'])}
  AND part IS {sql_str(row['part'])}
  AND subpart IS {sql_str(row['subpart'])}
ON CONFLICT(question_id, label) DO UPDATE SET
  answer_raw  = excluded.answer_raw,
  answer_code = excluded.answer_code,
  diagram_key = excluded.diagram_key;"""


def emit_diagram_upsert(d: dict) -> str:
    return (
        f"INSERT INTO diagrams (diagram_key, drive_path, status) "
        f"VALUES ({sql_str(d['diagram_key'])}, {sql_str(d['drive_path'])}, {sql_str(d['status'])}) "
        f"ON CONFLICT(diagram_key) DO UPDATE SET "
        f"drive_path = excluded.drive_path, status = excluded.status;"
    )


# =============================================================================
# Main
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description='CAPE Excel → D1 ETL')
    parser.add_argument('--p1',  required=True, help='Path to Cape_P1_s.xlsx')
    parser.add_argument('--p2',  required=True, help='Path to Cape_P2_s.xlsx')
    parser.add_argument('--out', default='load_questions.sql', help='Output SQL file')
    args = parser.parse_args()

    p1_path = Path(args.p1)
    p2_path = Path(args.p2)
    out_path = Path(args.out)

    for p in (p1_path, p2_path):
        if not p.exists():
            sys.exit(f"File not found: {p}")

    print(f"Loading P1 workbook: {p1_path}")
    wb_p1 = openpyxl.load_workbook(p1_path, data_only=True, read_only=True)
    print(f"Loading P2 workbook: {p2_path}")
    wb_p2 = openpyxl.load_workbook(p2_path, data_only=True, read_only=True)

    sql_lines: list[str] = [
        "-- Auto-generated by etl.py — do not edit by hand.",
        "-- Re-run etl.py to regenerate after spreadsheet changes.",
        "PRAGMA foreign_keys = ON;",
        "",
    ]

    stats = {
        'p1_questions': 0, 'p1_choices': 0,
        'p2_questions': 0, 'diagrams': 0,
        'skipped': 0,
    }

    seen_subjects: set[str] = set()

    def ensure_subject(subject: str):
        if subject and subject not in seen_subjects:
            sql_lines.append(emit_subject_upsert(subject))
            seen_subjects.add(subject)

    # ── Paper 1 sheets ────────────────────────────────────────────────────────
    print("\nProcessing Paper 1 sheets…")
    for sheet_name in wb_p1.sheetnames:
        if not sheet_name.startswith('Converted -'):
            print(f"  [SKIP] '{sheet_name}' (not a Converted sheet)")
            continue
        print(f"  Processing P1 sheet: {sheet_name}")
        ws = wb_p1[sheet_name]
        for row in rows_from_p1(ws, sheet_name):
            if not row['subject']:
                stats['skipped'] += 1
                continue
            ensure_subject(row['subject'])
            sql_lines.append(emit_question_upsert(row, 'P1'))
            for choice in row['choices']:
                sql_lines.append(emit_choice_upsert(row, choice, 'P1'))
                stats['p1_choices'] += 1
            stats['p1_questions'] += 1

    # ── Paper 2 sheets ────────────────────────────────────────────────────────
    print("\nProcessing Paper 2 sheets…")
    for sheet_name in wb_p2.sheetnames:
        if sheet_name.lower().startswith('diagram'):
            # Process Diagrams sheet separately below
            continue
        if not sheet_name.startswith('Converted -'):
            print(f"  [SKIP] '{sheet_name}' (not a Converted sheet)")
            continue
        print(f"  Processing P2 sheet: {sheet_name}")
        ws = wb_p2[sheet_name]
        for row in rows_from_p2(ws, sheet_name):
            if not row['subject']:
                stats['skipped'] += 1
                continue
            ensure_subject(row['subject'])
            sql_lines.append(emit_question_upsert(row, 'P2'))
            stats['p2_questions'] += 1

    # ── Diagrams sheet ────────────────────────────────────────────────────────
    print("\nProcessing Diagrams sheet…")
    for sheet_name in wb_p2.sheetnames:
        if sheet_name.lower().startswith('diagram'):
            ws = wb_p2[sheet_name]
            for d in rows_from_diagrams(ws, sheet_name):
                sql_lines.append(emit_diagram_upsert(d))
                stats['diagrams'] += 1
            print(f"  Parsed {stats['diagrams']} diagram rows from '{sheet_name}'")

    # ── Write output ──────────────────────────────────────────────────────────
    out_path.write_text('\n'.join(sql_lines), encoding='utf-8')

    print(f"\n{'='*60}")
    print(f"Output written to: {out_path}")
    print(f"  P1 questions : {stats['p1_questions']}")
    print(f"  P1 choices   : {stats['p1_choices']}")
    print(f"  P2 questions : {stats['p2_questions']}")
    print(f"  Diagrams     : {stats['diagrams']}")
    print(f"  Skipped rows : {stats['skipped']}")
    print(f"{'='*60}")
    print("\nNext steps:")
    print("  wrangler d1 execute cape-questions --local --file=load_questions.sql")
    print("  wrangler d1 execute cape-questions --file=load_questions.sql")


if __name__ == '__main__':
    main()
